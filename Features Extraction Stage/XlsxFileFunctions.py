#!/usr/bin/python
# -*- coding: utf-8 -*-
# author: Mohammad Modallal <m.modallal@redcrow.co> 
'''
This model used to read/write data from/to excel files
'''
from xlrd import open_workbook
import xlwt
import re
import sys, getopt, codecs, os
from itertools import izip
import csv
from openpyxl import Workbook

class DataInput(object):
    '''
       Class DataInput Description:
              This class represent lines that will read from .xlsx file.
       Fields:
              1)ArabicTextValue:
                                This field include Arabic string.   
              2)ArabicTextLabel:          
                                This field include Arabic text label. For example: Modern Standard Arabic, Gulf, Egyptian, Iraqi and Levantine
    '''
    def __init__(self, arabicTextValue, arabicTextLabel):
        self.arabicTextValue = ' '+arabicTextValue+' '
        self.arabicTextLabel = arabicTextLabel

    def __str__(self):
        return("DataInput object:\n"
               "  arabicTextValue = {0}\n"
               "  arabicTextLabel = {1}"
               .format(self.arabicTextValue, self.arabicTextLabel))

class TextFreq(object):
    '''
       Class TextFreq Description:
              This class represent Word_Freq that will store in .xlsx file.
       Fields:
              1)arabicWord:
                                This field include Arabic Word.   
              2)arabicFreq:          
                                This field include Arabic Frequency
    '''
    def __init__(self, arabicWord, arabicFreq):
        self.arabicWord = arabicWord
        self.arabicFreq = arabicFreq

    def __str__(self):
        return("DataInput object:\n"
               "  arabicWord = {0}\n"
               "  arabicFreq = {1}"
               .format(self.arabicWord, self.arabicFreq))

def ReadFile(fileName):
	'''
          Description:
                 ReadFile function used to read .xlsx files that contains Arabic sentences with labels
          Input:
                 1)fileName - File name with .Xlsx extension
          Output:
                 1) lines - List of DataInput objects that represent Arabic strings with labels
                 2) NofRowsList- List of number of rows in each sheet in the file.
                 3) NofColumnsList- List of number of columns in each sheet in the file.
        '''
        wb = open_workbook(fileName)                        #Open xlsx file which contains Arabic text with labels
	lines = []                                          #List of DataInput objects                           
	NofRowsList=[]                                      #List of sheets number_of_columns
        NofColumnsList=[]                                   #List of sheets number_of_rows
	for sheet in wb.sheets():                           #read file sheet by sheet
    		number_of_rows = sheet.nrows                #number of rows in sheet
                NofRowsList.append(number_of_rows)
    		number_of_columns = sheet.ncols             #number of columns in sheet
                NofColumnsList.append(number_of_columns)
    		rows = []
                for row in range(1, number_of_rows):
                   values = []
		   for col in range(number_of_columns):
                       value  = (sheet.cell(row,col).value)
                       try:
                          value = str(int(value))
                       except ValueError:
                          pass
                       finally:
                           values.append(value)
                   line = DataInput(*values)
                   lines.append(line)
        return lines,NofRowsList,NofColumnsList

def ReadFeaturesFile(fileName):
	'''
          Description:
                 ReadFeaturesFile function used to read .xlsx files that contains Arabic texts with frequencies
          Input:
                 1)fileName - File name with .Xlsx extension
          Output:
                 1) lines - List of TextFreq objects that represent Arabic texts with frequencies
                 2) NofRowsList- List of number of rows in each sheet in the file.
                 3) NofColumnsList- List of number of columns in each sheet in the file.
        '''
        wb = open_workbook(fileName)                        #Open xlsx file which contains Arabic text with labels
	lines = []                                          #List of DataInput objects                           
	NofRowsList=[]                                      #List of sheets number_of_columns
        NofColumnsList=[]                                   #List of sheets number_of_rows
	for sheet in wb.sheets():                           #read file sheet by sheet
    		number_of_rows = sheet.nrows                #number of rows in sheet
                NofRowsList.append(number_of_rows)
    		number_of_columns = sheet.ncols             #number of columns in sheet
                NofColumnsList.append(number_of_columns)
    		rows = []
                for row in range(1, number_of_rows):
                   values = []
		   for col in range(number_of_columns):
                       value  = (sheet.cell(row,col).value)
                       try:
                          value = str(int(value))
                       except ValueError:
                          pass
                       finally:
                           values.append(value)
                   line = TextFreq(*values)
                   lines.append(line)
        return lines,NofRowsList,NofColumnsList

def Write_Feature_Freq_ToFile(fileName,lines):
    '''
          Description:
                 Write_Feature_Freq_ToFile function used to write features with frequency to .xlsx files.
          Input:
                 1)fileName: File name with .Xlsx extension
                 2)lines: List of feature-freq.
    '''
    wb = Workbook()
    ws = wb.active
    # appended Rows
    for line in lines:
        ws.append([line.arabicWord, line.arabicFreq])
    wb.save(fileName)


def Write_FeaturesVector_ToFile(fileName,lines):
    '''
          Description:
                 Write_FeaturesVector_ToFile function used to write features to .xlsx files.
          Input:
                 1)fileName: File name with .Xlsx extension
                 2)lines: List of features.
    '''
    wb = Workbook()
    ws = wb.active
    # appended Rows
    for line in lines:
        ws.append([line])
    wb.save(fileName)

def WriteToFile(fileName,lines):
    '''
          Description:
                 WriteToFile function used to write arabicText with its Label to .xlsx files.
          Input:
                 1)fileName: File name with .Xlsx extension
                 2)lines: List of arabicText-Label.
    '''
    wb = Workbook()
    ws = wb.active
    # appended Rows
    for line in lines:
        ws.append([line.arabicTextValue, line.arabicTextLabel])
    wb.save(fileName)

