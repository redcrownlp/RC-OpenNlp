#!/usr/bin/python
# -*- coding: utf-8 -*-
# author: Mohammad Modallal <m.modallal@redcrow.co> 

import XlsxFileFunctions 
import nltk
from nltk import ngrams
from nltk.collocations import *
import collections
from nltk.corpus import stopwords
import re
import regex
import string
import warnings
warnings.simplefilter("ignore")
from numpy.core.umath_tests import inner1d
import numpy as np
import cPickle
import pyarabic.araby as araby
import pyarabic.number as number
from xlrd import open_workbook
import xlwt
import re
import csv
from openpyxl import Workbook
import CleanArabicText
##############################################################################################
def FindFeatureVectorForInputText(line,features_vector):
     '''
     fill feature vector for arabic text line: 
         if the feature exist in arabic text line then:
             append 1 to feature vector
         else: 
             append 0 to feature vector
     '''
     textValue=line
     textValue=' '+textValue+ ' '
     line_features=[]
     for word in features_vector:
         if word in textValue:
            line_features.append(1)  #line contains this feature 
         else:
            line_features.append(0)  #line not contains this feature 
     return line_features

    
########################################################################################## 
class TestingResult(object):
    '''
       Class TestingResult Description:
              This class represent lines that will read from .xlsx file.
       Fields:
              1)ArabicTextValue:
                                This field include Arabic string.   
              2)ArabicTextLabel:          
                                This field include Arabic text labels. For example: Maghreb, Gulf, Egyptian, Iraqi and Levant.
    '''
    def __init__(self, arabicTextValue, label1,label2):
        self.arabicTextValue = arabicTextValue
        self.label1 = label1
        self.label2 = label2

    def __str__(self):
        return("TestingResult object:\n"
               "  arabicTextValue = {0}\n"
               "  label1 = {1}\n"
               "  label2 = {2}"
               .format(self.arabicTextValue, self.label1,self.label2))
##############################################################################################
def Write_Reselt_acc_ToFile(fileName,lines):
    '''
          Description:
                 Write_Feature_Freq_ToFile function used to write features with frequency to .xlsx files.
          Input:
                 1)fileName: File name with .Xlsx extension
                 2)lines: List of feature-freq.
    '''
    wb = Workbook()
    ws = wb.active
    # appended Rows
    for line in lines:
        ws.append([line.arabicTextValue, line.arabicTextLabel1, line.Label1_acc, line.arabicTextLabel2, line.Label2_acc])
    wb.save(fileName)
########################################################################################## 
class DataInput(object):
    '''
       Class DataInput Description:
              This class represent lines that will read from .xlsx file.
       Fields:
              1)ArabicTextValue:
                                This field include Arabic string.   
    '''
    def __init__(self, arabicTextValue, arabicTextLabel):
        self.arabicTextValue = ' '+arabicTextValue+' '

    def __str__(self):
        return("DataInput object:\n"
               "  arabicTextValue = {0}\n"
               "  arabicTextLabel = {1}"
               .format(self.arabicTextValue))
########################################################################################## 
def ReadFile(fileName):
	'''
          Description:
                 ReadFile function used to read .xlsx files that contains Arabic sentences with labels.
          Input:
                 fileName - Input File name.
          Output:
                 lines - List of DataInput objects that represent Arabic strings with labels.

        '''
        wb = open_workbook(fileName)                        #Open xlsx file which contains Arabic text with labels
	lines = []                                          #List of DataInput objects                           
	NofRowsList=[]                                      #List of sheets number_of_columns
        NofColumnsList=[]                                   #List of sheets number_of_rows
	for sheet in wb.sheets():                           #read file sheet by sheet
    		number_of_rows = sheet.nrows                #number of rows in sheet
                NofRowsList.append(number_of_rows)
    		number_of_columns = sheet.ncols             #number of columns in sheet
                NofColumnsList.append(number_of_columns)
    		rows = []
                for row in range(1, number_of_rows):
                   values = []
		   for col in range(number_of_columns):
                       value  = (sheet.cell(row,col).value)
                       try:
                          value = str(int(value))
                       except ValueError:
                          pass
                       finally:
                           values.append(value)
                   line = values[0]
                   lines.append(line)
        return lines

###################################################################################################################################################################
if __name__ == "__main__":
           #####################################################################
           #Read Feature Vector
	   features_vector=XlsxFileFunctions.ReadFeaturesVectorFile('Features_Vector.xlsx')
	   map(unicode,features_vector)
	   #####################################################################
           testing_dataset=ReadFile("Test_dataset.xlsx")
           testing_result=[]
           count2=0
           for text in testing_dataset:
                #####################################################################
		#Remove non-Arabic text from the input text
		#text=unicode(text, 'utf-8')
		text=CleanArabicText.Arabic_Text_Processing(text)
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ####################################################                   Six Topics Togathers                     ############################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
		# Read NLP Model
		clf=[]
                #####################################################################
                with open('./Models/LogisticRegression_All_Topic_Model.pkl', 'rb') as fid:
		           clf = cPickle.load(fid)

		line_features=FindFeatureVectorForInputText(text,features_vector)
		result_label=clf.predict([line_features])
		classesResult=clf.predict_proba([line_features]) 
		labels=clf.classes_
                #[u'Culture' u'Economy' u'Health' u'Politics' u'Religion' u'Sport']
                Culture_Topic=classesResult[0][0]*100
                Economy_Topic=classesResult[0][1]*100
                Health_Topic=classesResult[0][2]*100
                Politics_Topic=classesResult[0][3]*100
                Religion_Topic=classesResult[0][4]*100
                Sport_Topic=classesResult[0][5]*100

                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ####################################################               Algorthim for Topics Identification             #########################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                ############################################################################################################################################################
                if True :
                        label1=""
                        label2=""
                        label1_acc=0
                        label2_acc=0 
                        result_not_sorted = [('Culture', Culture_Topic),
		                             ('Economy', Economy_Topic),
		                             ('Health', Health_Topic),
		                             ('Politics', Politics_Topic),
		                             ('Religion', Religion_Topic),
		                             ('Sport', Sport_Topic),
		                            ]
                        result_sorted=sorted(result_not_sorted, key=lambda s : s[1],reverse=True)
                        count=0
                        label1_acc=0
                        label2_acc=0
                        for key, value in result_sorted:
		                   count=count+1
		                   if count==1:
				        label1=key
				        label1_acc=value
		                   elif count==2:
				        label2=key
				        label2_acc=value
				   else:
		                        break;
                        if sum(line_features) > 0:
		                if label1_acc>30 and label2_acc>30:
		                         testing_result.append(TestingResult(text,label1,label2))
		                else:
		                         testing_result.append(TestingResult(text,label1,""))
                        else:
					 testing_result.append(TestingResult(text,"",""))
            
           wb = Workbook()
           ws = wb.active
           for line in testing_result:
              ws.append([line.arabicTextValue, line.label1, line.label2])
           wb.save("result.xlsx")
           

         
